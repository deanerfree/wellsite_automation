import openpyxl
from openpyxl import load_workbook
porosity_log = "WL_0504937-04-04-17-074-25W4-00_Geoforms.xlsm"
data =  {'por0': 40.5, 'por1': 50.0, 'por2': 88.0, 'por3': 369.0, 'por4': 555.0, 'por5': 291.0, 'por6': 8.5, 'por7': 0}
wb = load_workbook(porosity_log)
sheets = wb.sheetnames
count = 0

for iteration, sheet in enumerate(sheets):
  if (sheet == 'Porosity Log'):
    count = iteration

porosity_sheet = wb[sheets[count]]

def input_por_data (data):
  for value in porosity_sheet:
    print('test')
input_por_data(data)