import openpyxl
from openpyxl import load_workbook

def create_workbook(spd_sht, val):
  wb = load_workbook(spd_sht, data_only=val, keep_vba=True)
  return wb

def check_rows_used(wb):
  rows={"rows_used":[],"unused_rows":[]}
  for iteration, row in enumerate(wb.iter_rows(min_row=5, min_col=1, max_col=2, max_row=19, values_only=True)):  
    if (row[0] == f"Leg {iteration}" and row[1] > 0):
      rows["rows_used"].append("Leg " + str(iteration))
    if (row[0] == f"Leg {iteration}" and row[1] == 0):
      rows["unused_rows"].append("Leg " + str(iteration))
  return rows

def write_to_wb(sheet, used_sheet, data):
  for col, val in enumerate(data.values()):
    sheet.cell(row=int(used_sheet["rows_used"][-1][3:])+6, column=col+3).value = val